import time
import requests
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


def get_timestamp(timestr):  # 获取unix时间戳（毫秒）
    datetime_obj = datetime.strptime(timestr, "%Y-%m-%d %H:%M:%S")
    obj_stamp = int(time.mktime(datetime_obj.timetuple()) * 1000.0)
    return obj_stamp


def get_datetime(timeNum):  # 获取字符串时间戳（毫秒）
    timeStamp = float(timeNum / 1000)
    timeArray = time.localtime(timeStamp)
    timeStr = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
    return timeStr


def utc(local_datetime):
    local_unix = get_timestamp(local_datetime)
    utc_unix = local_unix - 28800000
    utc_datetime = str(get_datetime(utc_unix))
    return utc_datetime, utc_unix


def date_and_time(worksheet, r):
    try:
        date_time = worksheet.cell(row=r, column=1).value

        time = worksheet.cell(row=r, column=2).value
        hour = time.hour
        min = time.minute
        sec = time.second
        date_time = date_time.replace(hour=hour, minute=min, second=sec)

        return date_time

    except AttributeError:
        return worksheet.cell(row=r, column=1).value


def get_direction(worksheet, r):
    direction = worksheet.cell(row=r, column=5).value
    direction = direction.replace(' ', '')
    long_view = '看多'
    short_view = '看空'
    if direction == long_view:
        return True
    if direction == short_view:
        return False


def get_maxrow(worksheet):
    for i in range(1, worksheet.max_row + 2):
        while worksheet.cell(row=i + 1, column=5).value is None:
            return i
            break


def max_price(data3):
    list1 = []
    for i in range(len(data3)):
        list1.append(data3[i][2])
    price = max(list1)
    return price


def min_price(data3):
    list1 = []
    for i in range(len(data3)):
        list1.append(data3[i][3])
    price = min(list1)
    return price


def get_data(coin_type, utc_start_unix, end_unix, interval):  # 请求一天内每三分钟的数据
    BASE_URL = 'https://api.binance.com'
    Kline = '/api/v1/klines'
    limit = 1000
    kline_url = BASE_URL + Kline + '?' + 'symbol=' + coin_type + '&interval=' + interval + '&startTime=' + str(
        utc_start_unix) + '&endTime=' + str(end_unix) + '&limit=' + str(limit)
    resp = requests.get(kline_url)
    data = resp.json()
    return data


def stop_loss_1(data, stop_line, direction):  # 是否触发止损

    if direction:

        touch_index = 0

        for i in range(len(data)):

            point_time = data[i][0]
            point_price = data[i][1]

            if float(point_price) <= stop_line:
                touch_status = True
                touch_time = point_time
                touch_index = i
                touch_price = point_price
                break
            else:
                touch_status = False
    else:
        touch_index = 0

        for i in range(len(data)):

            point_time = data[i][0]
            point_price = data[i][1]

            if float(point_price) >= stop_line:
                touch_status = True
                touch_time = point_time
                touch_index = i
                touch_price = point_price
                break
            else:
                touch_status = False

    return touch_status, touch_index


def stop_loss_2(data, stop_percent, start_price, direction):
    if direction:

        touch_index = 0

        for i in range(len(data)):

            point_time = data[i][0]
            point_price = data[i][1]
            point_percent = (float(point_price) - start_price) / start_price

            if point_percent <= (-1) * stop_percent:
                touch_status = True
                touch_index = i
                break
            else:
                touch_status = False
    else:

        touch_index = 0

        for i in range(len(data)):

            point_time = data[i][0]
            point_price = data[i][1]
            point_percent = (start_price - float(point_price)) / start_price

            if point_percent <= (-1) * stop_percent:
                touch_status = True
                touch_index = i
                break
            else:
                touch_status = False

    return touch_status, touch_index


def liquidation(data, start_price, leverage, direction):
    if direction:
        liquidation_index = 0

        for i in range(len(data)):
            point_unix = data[i][0]
            point_price = data[i][1]
            point_loss = (float(point_price) - start_price) / start_price
            point_loss_after_leverage = point_loss * leverage
            if point_loss_after_leverage <= -1:
                liquidation_status = True
                liquidation_index = i
                break
            else:
                liquidation_status = False
    else:
        liquidation_index = 0

        for i in range(len(data)):
            point_unix = data[i][0]
            point_price = data[i][1]
            point_loss = (start_price - float(point_price)) / start_price
            point_loss_after_leverage = point_loss * leverage
            if point_loss_after_leverage <= -1:
                liquidation_status = True
                liquidation_index = i
                break
            else:
                liquidation_status = False

    return liquidation_status, liquidation_index


def play_time_2side(data, lowerbound, upperbound):  # 入场时间

    strike_index = 0

    for i in range(len(data)):

        point_price = data[i][1]
        point_index = i

        if (float(point_price) >= float(lowerbound)) and (float(point_price) <= float(upperbound)):
            strike_index = point_index
            play_or_not = True
            break
        else:
            play_or_not = False

    return play_or_not, strike_index


def play_time_1side(data, price_line, direction):
    strike_price = data[0][1]
    strike_index = 0

    if direction:

        for i in range(len(data)):

            point_time = data[i][0]
            point_price = data[i][1]
            point_index = i

            if float(point_price) < price_line:
                strike_price = point_price
                strike_index = point_index
                play_or_not = True
                break
            else:
                play_or_not = False
    else:

        for i in range(len(data)):

            point_price = data[i][1]
            point_index = i

            if float(point_price) > price_line:
                strike_price = point_price
                strike_index = point_index
                play_or_not = True
                break
            else:
                play_or_not = False

    return play_or_not, strike_index


def get_start_time_and_price(data_start_revised):
    start_price = float(data_start_revised[0][1])
    start_time = str(get_datetime(data_start_revised[0][0]))
    return start_time, start_price


def revise_start_position(worksheet, r, data_complete, lowerbound, upperbound):
    data_start_revised = data_complete
    play_status = False
    start_price = None

    if signal_type(worksheet, r) == 'current':
        data_start_revised = data_complete
        start_time, start_price = get_start_time_and_price(data_start_revised)
        worksheet.cell(row=r, column=17).value = start_time + '  ' + str(start_price)
        play_status = True

    elif signal_type(worksheet, r) == 'point':
        play_status, play_index = play_time_1side(data_complete, lowerbound, get_direction(worksheet, r))
        if play_status:
            data_start_revised = data_complete[play_index:]
            start_time, start_price = get_start_time_and_price(data_start_revised)
            worksheet.cell(row=r, column=17).value = start_time + '  ' + str(start_price)
        else:
            worksheet.cell(row=r, column=17).value = '未触发建议价格线'

    elif signal_type(worksheet, r) == 'interval':
        play_status, play_index = play_time_2side(data_complete, lowerbound, upperbound)
        if play_status:
            data_start_revised = data_complete[play_index:]
            start_time, start_price = get_start_time_and_price(data_start_revised)
            worksheet.cell(row=r, column=17).value = start_time + '  ' + str(start_price)
        else:
            worksheet.cell(row=r, column=17).value = '未触发建议价格区间'

    return data_start_revised, start_price, play_status


def revise_end_position(worksheet, r, data_start_revised, start_price, stop_line, stop_percent, leverage):
    # 检测是否止损
    stop_loss_index = len(data_start_revised)
    stop_loss_time = None
    stop_loss_price = None

    if stop_line is not None:
        stop_loss_status_1, stop_loss_index_1 = stop_loss_1(data_start_revised, stop_line, get_direction(worksheet, r))
        if stop_loss_status_1:
            stop_loss_time = data_start_revised[stop_loss_index_1][0]
            stop_loss_price = data_start_revised[stop_loss_index_1][1]
            stop_loss_index = stop_loss_index_1
        else:
            stop_loss_index = len(data_start_revised)
            stop_loss_time = '未触及止损线'
            stop_loss_price = None
    elif stop_percent is not None:
        stop_loss_status_2, stop_loss_index_2 = stop_loss_2(data_start_revised, stop_percent, start_price,
                                                            get_direction(worksheet, r))
        if stop_loss_status_2:
            stop_loss_time = data_start_revised[stop_loss_index_2][0]
            stop_loss_price = data_start_revised[stop_loss_index_2][1]
            stop_loss_index = stop_loss_index_2
        else:
            stop_loss_index = len(data_start_revised)
            stop_loss_time = '未触及止损线'
            stop_loss_price = None
    else:
        stop_loss_index = len(data_start_revised)

    # 检测是否爆仓
    liquidation_status, liquidation_index_1 = liquidation(data_start_revised, start_price, leverage,
                                                          get_direction(worksheet, r))
    liquidation_index = len(data_start_revised)
    liquidation_time = None
    liquidation_price = None
    if liquidation_status:
        liquidation_time = data_start_revised[liquidation_index_1][0]
        liquidation_price = data_start_revised[liquidation_index_1][1]
        liquidation_index = liquidation_index_1
    else:
        liquidation_index = len(data_start_revised)
        liquidation_time = '未爆仓'
        liquidation_price = None

    # 判断哪个index小，哪个小用哪个切片数据
    end_revised_index = len(data_start_revised)

    if stop_loss_index < liquidation_index:
        end_revised_index = stop_loss_index
        worksheet.cell(row=r, column=21).value = str(
            get_datetime(data_start_revised[stop_loss_index][0])) + '  ' + str(
            data_start_revised[stop_loss_index][1])
    elif stop_loss_index > liquidation_index:
        end_revised_index = liquidation_index
        worksheet.cell(row=r, column=22).value = str(
            get_datetime(data_start_revised[liquidation_index][0])) + '  ' + str(
            data_start_revised[liquidation_index][1])
    elif stop_loss_index == liquidation_index:
        end_revised_index = stop_loss_index

    data_end_revised = data_start_revised[:end_revised_index + 1]

    return data_end_revised


def data_analysis(filename):
    workbook = openpyxl.load_workbook(filename)

    for item in workbook.sheetnames:
        worksheet = workbook[item]
        maxrow = get_maxrow(worksheet)
        for i in range(1, maxrow):
            fill_in(worksheet, i + 1, signal_type(worksheet, i + 1))

    workbook.save(filename=filename)


def signal_type(worksheet, r):
    if worksheet.cell(row=r, column=6).value is None and worksheet.cell(row=r,
                                                                        column=7).value is None:
        signalType = 'current'
    elif worksheet.cell(row=r, column=6).value is not None and worksheet.cell(row=r,
                                                                              column=7).value is not None:
        signalType = 'interval'
    elif worksheet.cell(row=r, column=6).value is not None and worksheet.cell(row=r,
                                                                              column=7).value is None:
        signalType = 'point'

    return signalType


def calculate_and_fill_in(worksheet, r, data_end_revised, start_price, leverage, direction):
    return_sum = 0
    return_list = []
    max_ret_list = []
    if direction:
        for i in range(len(data_end_revised)):
            return_every_5min = (float(data_end_revised[i][1]) - start_price) / start_price
            max_ret_every_5min = (float(data_end_revised[i][2]) - start_price) / start_price
            return_list.append(return_every_5min)
            max_ret_list.append(max_ret_every_5min)
            return_sum = return_sum + return_every_5min
    else:
        for i in range(len(data_end_revised)):
            return_every_5min = (start_price - float(data_end_revised[i][1])) / start_price
            max_ret_every_5min = (start_price - float(data_end_revised[i][2])) / start_price
            return_list.append(return_every_5min)
            max_ret_list.append(max_ret_every_5min)
            return_sum = return_sum + return_every_5min

    average_return = return_sum / (len(data_end_revised))


    max_return = max(max_ret_list)

    sd_sum = 0
    sd_list = []
    for ret in return_list:
        sd_every_5min = (ret - average_return) ** 2
        sd_list.append(sd_every_5min)
        sd_sum = sd_sum + sd_every_5min

    volatility = ((sd_sum) / (len(data_end_revised) - 1)) ** (0.5)

    return average_return, max_return, volatility


def fill_in(worksheet, r, signalType):
    try:
        print('当前进行至：'+ str(worksheet) + str(r))
        # 获取时间
        local_start_datetime = str(date_and_time(worksheet, r))
        utc_start_datetime, utc_start_unix = utc(local_start_datetime)

        end_unix_24h = utc_start_unix + 86400000  # unix时间 +1d (毫秒)
        end_unix_72h = utc_start_unix + 86400000 * 3

        # 获取币种
        coin_type = worksheet.cell(row=r, column=3).value
        coin_type = coin_type.replace('/', '')

        # 获取区间
        lowerbound = worksheet.cell(row=r, column=6).value
        upperbound = worksheet.cell(row=r, column=7).value

        # 获取止损线
        stop_line = worksheet.cell(row=r, column=8).value
        stop_percent = worksheet.cell(row=r, column=9).value

        # 获取杠杆
        leverage = worksheet.cell(row=r, column=10).value
        if leverage is None:
            leverage = 1

        # 获取72h数据
        data_72h_complete = get_data(coin_type, utc_start_unix, end_unix_72h, '5m')
        # 切片出24h数据
        data_24h_complete = get_data(coin_type, utc_start_unix, end_unix_24h, '5m')
        # data_24h_complete = data_72h_complete[0:289]
        # 此处以上内容已检测，运行正常

        # 获取入场价格和入场时间，并切片数据
        data_start_revised_24h, start_price_24h, play_status_24h = revise_start_position(worksheet, r,
                                                                                         data_24h_complete, lowerbound,
                                                                                         upperbound)
        data_start_revised_72h, start_price_72h, play_status_72h = revise_start_position(worksheet, r,
                                                                                         data_72h_complete, lowerbound,
                                                                                         upperbound)

        # 如果24h入场，进行填写
        if play_status_24h:
            # 填写入场时间和价格
            worksheet.cell(row=r, column=17).value = str(get_datetime(data_start_revised_24h[0][0])) + '  ' + str(
                start_price_24h)
            # 检测止损和爆仓并填写止损或爆仓状态，获得切片数据，填写出场时间和价格
            data_end_revised_24h = revise_end_position(worksheet, r, data_start_revised_24h, start_price_24h, stop_line,
                                                       stop_percent, leverage)
            worksheet.cell(row=r, column=18).value = str(get_datetime(data_end_revised_24h[-1][0])) + '  ' + str(
                data_end_revised_24h[-1][1])
            # 计算return和volatility并填写
            average_return_24h, max_return_24h, volatility_24h = calculate_and_fill_in(worksheet, r,
                                                                                       data_end_revised_24h,
                                                                                       start_price_24h, leverage,
                                                                                       get_direction(worksheet, r))
            worksheet.cell(row=r, column=11).value = average_return_24h*leverage
            worksheet.cell(row=r, column=12).value = max_return_24h*leverage
            worksheet.cell(row=r, column=13).value = volatility_24h*leverage
            worksheet.cell(row=r, column=11).number_format = '0.00%'
            worksheet.cell(row=r, column=12).number_format = '0.00%'
            worksheet.cell(row=r, column=13).number_format = '0.00%'
            fill = PatternFill("solid", fgColor="FAEBD7")
            worksheet.cell(row=r, column=11).fill = fill
            worksheet.cell(row=r, column=12).fill = fill
            worksheet.cell(row=r, column=13).fill = fill

        # 如果72h入场，进行填写
        if play_status_72h:
            # 填写入场时间和价格
            worksheet.cell(row=r, column=19).value = str(get_datetime(data_start_revised_72h[0][0])) + '  ' + str(
                start_price_72h)
            # 检测止损和爆仓并填写止损或爆仓状态，获得切片数据，填写出场时间和价格
            data_end_revised_72h = revise_end_position(worksheet, r, data_start_revised_72h, start_price_72h, stop_line,
                                                       stop_percent, leverage)
            worksheet.cell(row=r, column=20).value = str(get_datetime(data_end_revised_72h[-1][0])) + '  ' + str(
                data_end_revised_72h[-1][1])
            # 计算return和volatility并填写
            average_return_72h, max_return_72h, volatility_72h = calculate_and_fill_in(worksheet, r,
                                                                                       data_end_revised_72h,
                                                                                       start_price_72h, leverage,
                                                                                       get_direction(worksheet, r))
            worksheet.cell(row=r, column=14).value = average_return_72h*leverage
            worksheet.cell(row=r, column=15).value = max_return_72h*leverage
            worksheet.cell(row=r, column=16).value = volatility_72h*leverage
            worksheet.cell(row=r, column=14).number_format = '0.00%'
            worksheet.cell(row=r, column=15).number_format = '0.00%'
            worksheet.cell(row=r, column=16).number_format = '0.00%'
            fill = PatternFill("solid", fgColor="F0F8FF")
            worksheet.cell(row=r, column=14).fill = fill
            worksheet.cell(row=r, column=15).fill = fill
            worksheet.cell(row=r, column=16).fill = fill


    except KeyError:
        print(worksheet, '第' + str(r) + "行数据，交易对不合法")
    except ZeroDivisionError:
        print(worksheet, '第' + str(r) + "行数据")


filename = r'信号追踪.xlsx'
data_analysis(filename)

