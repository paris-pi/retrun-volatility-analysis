import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import time
from datetime import datetime

font_title = Font(u'等线', size=11, bold=True, italic=False, strike=False, color='000000')

def exsisted_sheet(sheetname,workbook):
    if sheetname in workbook.sheetnames:
        return True
    else:
        return False

def date_and_time(worksheet, r):
    try:
        date_time = worksheet.cell(row=r, column=1).value

        time = worksheet.cell(row=r, column=2).value
        hour = time.hour
        min = time.minute
        sec = time.second
        date_time = date_time.replace(hour=hour, minute=min, second=sec)

        return date_time

    except AttributeError:
        return worksheet.cell(row=r, column=1).value

def get_maxrow(worksheet):
    for i in range(1, worksheet.max_row + 2):
        while worksheet.cell(row=i + 1, column=1).value is None:
            return i
            break

def have_data(worksheet,r):
    if worksheet.cell(row=r,column=14).value is not None:
        return True
    else: return  False

def adjust_width(worksheet):
    worksheet.column_dimensions['A'].width = 18
    worksheet.column_dimensions['B'].width = 10.67

def adjust_width_with_source(worksheet):
    worksheet.column_dimensions['A'].width = 8
    worksheet.column_dimensions['B'].width = 18

def fill_data(ws_in,ws_base,r_in,r_base):
    ws_base.cell(row=r_base, column=1).value = date_and_time(ws_in,r_in)
    for i in range(1,22):
        ws_base.cell(row=r_base, column=i+1).value = ws_in.cell(row=r_in, column=i+1).value

def fill_data_with_source(ws_in,ws_base,r_in,r_base):
    ws_base.cell(row=r_base, column=2).value = date_and_time(ws_in,r_in)
    for i in range(1,22):
        ws_base.cell(row=r_base, column=i+2).value = ws_in.cell(row=r_in, column=i+1).value

def pattern_fill(worksheet,r):
    worksheet.cell(row=r, column=11).number_format = '0.00%'
    worksheet.cell(row=r, column=12).number_format = '0.00%'
    worksheet.cell(row=r, column=13).number_format = '0.00%'
    fill = PatternFill("solid", fgColor="FAEBD7")
    worksheet.cell(row=r, column=11).fill = fill
    worksheet.cell(row=r, column=12).fill = fill
    worksheet.cell(row=r, column=13).fill = fill
    worksheet.cell(row=r, column=14).number_format = '0.00%'
    worksheet.cell(row=r, column=15).number_format = '0.00%'
    worksheet.cell(row=r, column=16).number_format = '0.00%'
    fill = PatternFill("solid", fgColor="F0F8FF")
    worksheet.cell(row=r, column=14).fill = fill
    worksheet.cell(row=r, column=15).fill = fill
    worksheet.cell(row=r, column=16).fill = fill

def pattern_fill_one_sheet(worksheet,r):
    worksheet.cell(row=r, column=12).number_format = '0.00%'
    worksheet.cell(row=r, column=13).number_format = '0.00%'
    worksheet.cell(row=r, column=14).number_format = '0.00%'
    fill = PatternFill("solid", fgColor="FAEBD7")
    worksheet.cell(row=r, column=12).fill = fill
    worksheet.cell(row=r, column=13).fill = fill
    worksheet.cell(row=r, column=14).fill = fill
    worksheet.cell(row=r, column=15).number_format = '0.00%'
    worksheet.cell(row=r, column=16).number_format = '0.00%'
    worksheet.cell(row=r, column=17).number_format = '0.00%'
    fill = PatternFill("solid", fgColor="F0F8FF")
    worksheet.cell(row=r, column=15).fill = fill
    worksheet.cell(row=r, column=16).fill = fill
    worksheet.cell(row=r, column=17).fill = fill

def save_in_different_worksheets(filename_in,filename_base):

    title = ['日期(GMT+8)','时间', '交易对','期限','多空','建议入场价下限（或建议入场价）','建议入场价上限', '止损线（价格）','止损线（百分比）','杠杆倍数','24H平均收益','24H最大收益','24H收益波动性','72H平均收益','72H最大收益','72H收益波动性','24H入场状态','24H出场状态','72H入场状态','72H出场状态','止损时间及价格','爆仓时间及价格']

    wb_in = openpyxl.load_workbook(filename_in)
    wb_base = openpyxl.load_workbook(filename_base)

    for name in wb_in.sheetnames:

        ws_in = wb_in[name]
        maxrow_in = get_maxrow(ws_in)

        if exsisted_sheet(name,wb_base):
            ws_base = wb_base[name]
            maxrow_base = get_maxrow(ws_base)
            blank_list = []
            for i in range(1,maxrow_in):
                if have_data(ws_in,i+1):
                    fill_data(ws_in,ws_base,i+1,maxrow_base+i)
                    pattern_fill(ws_base,maxrow_base+i)
                else: blank_list.append(maxrow_base+i)
            for j in range(len(blank_list)-1,-1,-1):
                ws_base.delete_rows(blank_list[j])

        else:
            ws_base = wb_base.create_sheet(name)
            adjust_width(ws_base)
            blank_list = []
            for k in range(len(title)):
                ws_base.cell(row=1,column=k+1).value = title[k]
                ws_base.cell(row=1, column=k+1).font = font_title
            for i in range(1,maxrow_in):
                if have_data(ws_in,i+1):
                    fill_data(ws_in,ws_base,i+1,i+1)
                    pattern_fill(ws_base,i+1)
                else: blank_list.append(i+1)
            for j in range(len(blank_list)-1,-1,-1):
                ws_base.delete_rows(blank_list[j])


    wb_base.save(filename=filename_base)


def save_in_one_worksheet(filename_in,filename_base):

    wb_in = openpyxl.load_workbook(filename_in)
    wb_base = openpyxl.load_workbook(filename_base)
    ws_base = wb_base.active
    adjust_width_with_source(ws_base)

    title = ['信号源','日期(GMT+8)','时间', '交易对','期限','多空','建议入场价下限（或建议入场价）','建议入场价上限', '止损线（价格）','止损线（百分比）','杠杆倍数','24H平均收益','24H最大收益','24H收益波动性','72H平均收益','72H最大收益','72H收益波动性','24H入场状态','24H出场状态','72H入场状态','72H出场状态','止损时间及价格','爆仓时间及价格']

    for k in range(len(title)):
        ws_base.cell(row=1, column=k + 1).value = title[k]
        ws_base.cell(row=1, column=k + 1).font = font_title

    for name in wb_in.sheetnames:

        ws_in = wb_in[name]
        maxrow_in = get_maxrow(ws_in)
        maxrow_base = get_maxrow(ws_base)
        blank_list = []

        for i in range(1, maxrow_in):
            if have_data(ws_in, i + 1):
                ws_base.cell(row=maxrow_base+i, column=1).value = name
                fill_data_with_source(ws_in,ws_base,i+1,maxrow_base+i)
                pattern_fill_one_sheet(ws_base,maxrow_base+i)
            else:
                blank_list.append(maxrow_base + i)
        for j in range(len(blank_list) - 1, -1, -1):
            ws_base.delete_rows(blank_list[j])

    wb_base.save(filename=filename_base)

def data_summary(worksheet):

    maxrow = get_maxrow(worksheet)
    print(worksheet)
    print('信号数：' + str(maxrow-1))
    #24h

    #ret
    ret_24h_list = []
    for i in range(1,maxrow):
        if worksheet.cell(row=i+1,column=11).value is not None:
            ret_24h_list.append(worksheet.cell(row=i+1,column=11).value)
    ret_average_24h = sum(ret_24h_list)/len(ret_24h_list)
    print('24h平均”平均收益“是：{:.2%}'.format(ret_average_24h))

    # max
    maxret_24h_list = []
    for i in range(1,maxrow):
        if worksheet.cell(row=i + 1, column=12).value is not None:
            maxret_24h_list.append(worksheet.cell(row=i+1,column=12).value)
    maxret_average_24h = sum(maxret_24h_list)/len(maxret_24h_list)
    print('24h平均“最大收益”是：{:.2%}'.format(maxret_average_24h))

    # vol
    vol_24h_list = []
    for i in range(1,maxrow):
        if worksheet.cell(row=i + 1, column=13).value is not None:
            vol_24h_list.append(worksheet.cell(row=i+1,column=13).value)
    vol_average_24h = sum(vol_24h_list)/len(vol_24h_list)
    print('24h平均“波动性”是：{:.2%}'.format(vol_average_24h))

    # 72h

    # ret
    ret_72h_list = []
    for i in range(1, maxrow):
        if worksheet.cell(row=i + 1, column=14).value is not None:
            ret_72h_list.append(worksheet.cell(row=i + 1, column=14).value)
    ret_average_72h = sum(ret_72h_list) / len(ret_72h_list)
    print('72h平均”平均收益“是：{:.2%}'.format(ret_average_72h))

    # max
    maxret_72h_list = []
    for i in range(1, maxrow):
        if worksheet.cell(row=i + 1, column=15).value is not None:
            maxret_72h_list.append(worksheet.cell(row=i + 1, column=15).value)
    maxret_average_72h = sum(maxret_72h_list) / len(maxret_72h_list)
    print('72h平均“最大收益”是：{:.2%}'.format(maxret_average_72h))

    # vol
    vol_72h_list = []
    for i in range(1, maxrow):
        if worksheet.cell(row=i + 1, column=16).value is not None:
            vol_72h_list.append(worksheet.cell(row=i + 1, column=16).value)
    vol_average_72h = sum(vol_72h_list) / len(vol_72h_list)
    print('72h平均“波动性”是：{:.2%}'.format(vol_average_72h))


#
file_list = [r"C:\Users\shang\OneDrive\Desktop\Paris工作包\主流币追踪\return_volatility_analysis\信号追踪5.xlsx"]
filename_base = r'C:\Users\shang\OneDrive\Desktop\data_in_one_sheet.xlsx'

for file in file_list:
    save_in_one_worksheet(file,filename_base)


# 对database文件使用，生成summary
# filename = r"C:\Users\shang\OneDrive\Desktop\Paris工作包\主流币追踪\database\data_in_different_sheets.xlsx"
# workbook = openpyxl.load_workbook(filename)
# for name in workbook.sheetnames:
#     worksheet = workbook[name]
#     data_summary(worksheet)
#     print('------------------')
