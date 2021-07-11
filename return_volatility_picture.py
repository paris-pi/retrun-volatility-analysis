import openpyxl
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
import numpy as np
from matplotlib import cm
import matplotlib.ticker as mtick
from os import listdir
import os
import PIL.Image as Image
from matplotlib.ticker import FuncFormatter
import numpy as np


def get_maxrow(worksheet):
    for i in range(1, worksheet.max_row + 2):
        while worksheet.cell(row=i + 1, column=3).value is None:
            return i
            break


def have_24h_data(worksheet, r):
    if worksheet.cell(row=r, column=11).value is not None:
        return True
    else:
        return False


def have_72h_data(worksheet, r):
    if worksheet.cell(row=r, column=14).value is not None:
        return True
    else:
        return False


def get_axis_data(workbook,sheetname, duration):

    worksheet = workbook[sheetname]
    maxrow = get_maxrow(worksheet)

    if duration == '24h':

        average_ret_24h = []
        max_ret_24h = []
        vol_24h = []
        coinpair_list = []
        signal_No = []

        for i in range(1, maxrow):
            if have_24h_data(worksheet, i + 1):
                average_ret = worksheet.cell(row=i + 1, column=11).value
                max_ret = worksheet.cell(row=i + 1, column=12).value
                vol = worksheet.cell(row=i + 1, column=13).value
                coinpair = worksheet.cell(row=i + 1, column=3).value
                signal_No.append(str(i))
                coinpair_list.append(coinpair)
                average_ret_24h.append(average_ret)
                max_ret_24h.append(max_ret)
                vol_24h.append(vol)

        return signal_No, coinpair_list, average_ret_24h, max_ret_24h, vol_24h

    if duration == '72h':

        average_ret_72h = []
        max_ret_72h = []
        vol_72h = []
        coinpair_list = []
        signal_No = []

        for i in range(1, maxrow):
            if have_72h_data(worksheet, i + 1):
                average_ret = worksheet.cell(row=i + 1, column=14).value
                max_ret = worksheet.cell(row=i + 1, column=15).value
                vol = worksheet.cell(row=i + 1, column=16).value
                coinpair = worksheet.cell(row=i + 1, column=3).value
                signal_No.append(str(i))
                coinpair_list.append(coinpair)
                average_ret_72h.append(average_ret)
                max_ret_72h.append(max_ret)
                vol_72h.append(vol)

        return signal_No, coinpair_list, average_ret_72h, max_ret_72h ,vol_72h

def to_percent(temp, position):
    return '%1.0f' % (100 * temp) + '%'

def plot_chart(sheetname,signal_No,coinpair_list,average_ret_list,max_ret_list,vol_list,duration):


    fig = plt.figure(figsize=(15, 10), dpi=80,facecolor='black')
    plt.style.use("dark_background")

    ax = fig.add_subplot(111)
    plt.rcParams['font.sans-serif']=['SimHei']  #显示中文标签
    plt.rcParams['axes.unicode_minus']=False

    # 构建24h数据
    x_data = signal_No
    y_data = average_ret_list
    y_data2 = max_ret_list
    y_data3 = vol_list

    # 设置正负柱显示不同颜色
    colors = []
    for i in y_data2:
        if i >= 0:
            colors.append('#05B98C')
        if i < 0:
            colors.append('#CA3950')

    # 绘图
    # ax.scatter(x_data, y_data, marker='_', s=1000, c='white', label='平均收益')
    ax.bar(x=x_data, height=y_data2, label='最大收益',width=0.6, color=colors)
    ax2 = ax.twinx()
    ax2.plot(x_data, y_data3, marker='.', c='#CA3950', label='波动性')
    ax.set_ylim((-0.1,0.5))
    ax2.set_ylim((-0.1,0.2))
    # 设置刻度颜色
    plt.rcParams['xtick.color'] = 'white'
    plt.rcParams['ytick.color'] = 'white'
    ax.yaxis.set_major_formatter(FuncFormatter(to_percent))
    ax2.yaxis.set_major_formatter(FuncFormatter(to_percent))

    # 在柱状图上显示具体数值, ha参数控制水平对齐方式, va控制垂直对齐方式

    for x, y, z in zip(x_data,y_data2,coinpair_list):
        ax.text(x, y+0.01 , '%s' % z, ha='center', va='bottom',color='white',fontsize=12, rotation=90)

    # 设置标题

    # 为两条坐标轴设置名称
    ax.set_xlabel("信号",fontsize=20,color='white')
    ax.set_ylabel("收益率",fontsize=20,color='white')
    ax2.set_ylabel("波动性",fontsize=20,color='white')

    # 显示图例
    ax.legend(loc='lower left')
    ax2.legend(loc='lower right')

    if duration == '24h':
        ax.patch.set_facecolor("black")
        plt.title(sheetname+' 24h预测表现', fontsize=40,color='white')
        plt.savefig('C:\\Users\shang\OneDrive\Desktop\可视化\子图\\' + sheetname + '\\' + sheetname + '24h预测表现' + '.png')
        plt.savefig('C:\\Users\shang\OneDrive\Desktop\可视化\所有子图\\'  + sheetname + '24h预测表现' + '.png')
    if duration == '72h':
        ax.patch.set_facecolor("black")
        plt.title(sheetname + ' 72h预测表现', fontsize=40,color='white')
        plt.savefig('C:\\Users\shang\OneDrive\Desktop\可视化\子图\\' + sheetname + '\\' + sheetname + '72h预测表现' + '.png')
        plt.savefig('C:\\Users\shang\OneDrive\Desktop\可视化\所有子图\\' + sheetname + '72h预测表现' + '.png')
    # plt.show()





# 定义图像拼接函数
def image_compose(sheetname):

    IMAGES_PATH = 'C:\\Users\shang\OneDrive\Desktop\可视化\子图\\' + sheetname + '\\' # 图片集地址
    IMAGES_FORMAT = ['.png']  # 图片格式
    IMAGE_SIZE = 1024  # 每张小图片的大小
    IMAGE_ROW = 1  # 图片间隔，也就是合并成一张图后，一共有几行
    IMAGE_COLUMN = 2  # 图片间隔，也就是合并成一张图后，一共有几列
    IMAGE_SAVE_PATH = 'C:\\Users\shang\OneDrive\Desktop\可视化\拼接合成图\\' + sheetname + '.png'  # 图片转换后的地址

    # 获取图片集地址下的所有图片名称
    image_names = [name for name in os.listdir(IMAGES_PATH) for item in IMAGES_FORMAT if
                   os.path.splitext(name)[1] == item]

    # 简单的对于参数的设定和实际图片集的大小进行数量判断
    if len(image_names) != IMAGE_ROW * IMAGE_COLUMN:
        raise ValueError("合成图片的参数和要求的数量不能匹配！")

    to_image = Image.new('RGB', (IMAGE_COLUMN * IMAGE_SIZE, IMAGE_ROW * IMAGE_SIZE))  # 创建一个新图
    # 循环遍历，把每张图片按顺序粘贴到对应位置上
    for y in range(1, IMAGE_ROW + 1):
        for x in range(1, IMAGE_COLUMN + 1):
            from_image = Image.open(IMAGES_PATH + image_names[IMAGE_COLUMN * (y - 1) + x - 1]).resize(
                (IMAGE_SIZE, IMAGE_SIZE), Image.ANTIALIAS)
            to_image.paste(from_image, ((x - 1) * IMAGE_SIZE, (y - 1) * IMAGE_SIZE))
    return to_image.save(IMAGE_SAVE_PATH)  # 保存新图




def mkdir(path):
    folder = os.path.exists(path)
    if not folder:  # 判断是否存在文件夹如果不存在则创建为文件夹
        os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
        print
        "---  new folder...  ---"
        print
        "---  OK  ---"
    else:
        print
        "---  There is this folder!  ---"

def action(workbook,sheetname):

    path = 'C:\\Users\shang\OneDrive\Desktop\可视化\子图\\' + sheetname + '\\'
    mkdir(path)

    signal_No, coinpair_list, average_ret, max_ret, vol = get_axis_data(workbook,sheetname,'24h')
    plot_chart(sheetname,signal_No,coinpair_list,average_ret,max_ret,vol,'24h')
    signal_No, coinpair_list, average_ret, max_ret, vol = get_axis_data(workbook,sheetname, '72h')
    plot_chart(sheetname, signal_No, coinpair_list, average_ret, max_ret, vol, '72h')

    image_compose(sheetname)



def weekly_picture(row):

    IMAGES_PATH = 'C:\\Users\shang\OneDrive\Desktop\可视化\所有子图\\' # 图片集地址
    IMAGES_FORMAT = ['.png']  # 图片格式
    IMAGE_SIZE = 1024  # 每张小图片的大小
    IMAGE_ROW = row  # 图片间隔，也就是合并成一张图后，一共有几行 (信息源个数)
    IMAGE_COLUMN = 2  # 图片间隔，也就是合并成一张图后，一共有几列 (每个信息源两张图，24h和72h)
    IMAGE_SAVE_PATH = 'C:\\Users\shang\OneDrive\Desktop\可视化\周报图\\' + '周报图' + '.png'  # 图片转换后的地址

    # 获取图片集地址下的所有图片名称
    image_names = [name for name in os.listdir(IMAGES_PATH) for item in IMAGES_FORMAT if
                   os.path.splitext(name)[1] == item]


    # 简单的对于参数的设定和实际图片集的大小进行数量判断
    if len(image_names) != IMAGE_ROW * IMAGE_COLUMN:
        raise ValueError("合成图片的参数和要求的数量不能匹配！")

    to_image = Image.new('RGB', (IMAGE_COLUMN * IMAGE_SIZE, IMAGE_ROW * IMAGE_SIZE))  # 创建一个新图
    # 循环遍历，把每张图片按顺序粘贴到对应位置上
    for y in range(1, IMAGE_ROW + 1):
        for x in range(1, IMAGE_COLUMN + 1):
            from_image = Image.open(IMAGES_PATH + image_names[IMAGE_COLUMN * (y - 1) + x - 1]).resize(
                (IMAGE_SIZE, IMAGE_SIZE), Image.ANTIALIAS)
            to_image.paste(from_image, ((x - 1) * IMAGE_SIZE, (y - 1) * IMAGE_SIZE))
    return to_image.save(IMAGE_SAVE_PATH)  # 保存新图



#生成给定数据的图片
def generate_picture(filename):
# filename = r'C:\Users\shang\OneDrive\Desktop\Paris工作包\主流币追踪\return_volatility_analysis\信号追踪.xlsx'
    workbook = openpyxl.load_workbook(filename)

    for sheetname in workbook.sheetnames:
        action(workbook,sheetname)

# 在一张图里生成所有数据源的所有图片
def all_imformation_picture(filename):
# filename = r'C:\Users\shang\OneDrive\Desktop\Paris工作包\主流币追踪\database\data_in_different_sheets.xlsx'
    workbook = openpyxl.load_workbook(filename)

    for sheetname in workbook.sheetnames:
        action(workbook,sheetname)

    weekly_picture(12)


# filename = r'C:\Users\shang\OneDrive\Desktop\Paris工作包\主流币追踪\database\data_in_different_sheets.xlsx'
# workbook = openpyxl.load_workbook(filename)
#
# action(workbook, 'NH')


filename = r'C:\Users\shang\OneDrive\Desktop\Paris工作包\主流币追踪\database\data_in_different_sheets.xlsx'
all_imformation_picture(filename)
